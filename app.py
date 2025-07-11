import os
import sqlite3
import pandas as pd
from datetime import datetime
from flask import Flask, request, render_template, redirect, url_for, jsonify, send_file, flash
from werkzeug.utils import secure_filename
import hashlib
import traceback
import io
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
import re
import gspread
from oauth2client.service_account import ServiceAccountCredentials
import json
from email.mime.text import MIMEText
import smtplib
from dotenv import load_dotenv
from email.mime.text import MIMEText
import smtplib

load_dotenv()


app = Flask(__name__)
app.secret_key = "your-secret-key"
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['DATABASE'] = 'database.db'
app.config['ALLOWED_EXTENSIONS'] = {'csv'}

os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

# Googleスプレッドシートへの接続

def connect_sheets():
    scope = ['https://spreadsheets.google.com/feeds', 'https://www.googleapis.com/auth/drive']

    json_path = os.environ.get('GOOGLE_CREDENTIALS_PATH')
    if not json_path:
        raise ValueError("⚠️ GOOGLE_CREDENTIALS_PATH が読み込めていません！")

    try:
        with open(json_path) as f:
            creds_dict = json.load(f)
    except Exception as e:
        raise ValueError(f"⚠️ JSONファイルの読み込みに失敗しました: {e}")

    creds = ServiceAccountCredentials.from_json_keyfile_dict(creds_dict, scope)
    client = gspread.authorize(creds)
    sheet = client.open('【開発用】シードル出庫台帳')
    return sheet.worksheet('集計用'),  sheet.worksheet('出庫情報'), sheet.worksheet('出庫詳細')


# --- メール送信用関数群 ---
def get_available_items():
    集計用シート, _, _ = connect_sheets()
    data = 集計用シート.get_all_records()
    available = [row for row in data if isinstance(row['現在庫'], int) and row['現在庫'] > 0]
    return available

def create_email_body(items):
    header = (
        "取引先様　各位\n"
        "いつもお世話になっております。有限会社マルカメ果樹園（マルカメ醸造所）の北沢毅です。\n"
        "現在のシードル/ワイン在庫をお送りいたします。\n"
        "発注の際はこちらのメール宛にご返信いただけましたら幸いです。\n\n"
        "【現在の在庫】\n"
    )
    if not items:
        stock_info = "現在、在庫のある商品はありません。\n"
    else:
        stock_info = "\n".join([f"- {row['商品名']}（{row['現在庫']}本）" for row in items]) + "\n"
    footer = "\nご不明点がございましたらお気軽にお問い合わせください。\n今後ともよろしくお願いいたします。\n"
    return header + stock_info + footer

def extract_valid_emails(raw_text):
    cleaned = re.sub(r'[\s\u3000]+', ',', raw_text)
    candidates = cleaned.split(',')
    pattern = r'^[\w\.-]+@[\w\.-]+\.\w+$'
    return [email.strip() for email in candidates if re.match(pattern, email.strip())]

def send_email(body, to_email):
    from_email = os.environ.get('EMAIL_USER')
    password = os.environ.get('EMAIL_PASS')
    msg = MIMEText(body)
    msg['Subject'] = "在庫情報のお知らせ"
    msg['From'] = from_email
    msg['To'] = to_email
    with smtplib.SMTP_SSL('smtp.gmail.com', 465) as server:
        server.login(from_email, password)
        server.send_message(msg)

def save_email_log_xlsx(email, status, error_msg=''):
    log_file = 'email_log.xlsx'
    if os.path.exists(log_file):
        wb = load_workbook(log_file)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "送信ログ"
        ws.append(['送信先', '送信日時', '結果', 'エラーメッセージ'])
    ws.append([
        email,
        datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        status,
        error_msg
    ])
    for column_cells in ws.columns:
        length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = length + 2
    wb.save(log_file)


#出庫詳細データ振り分け用の関数
def 振り分け処理(row, 出庫情報シート, sheets):
    出庫ID, 商品名, 数量 = row

    # 出庫IDから出庫先を取得
    出庫情報一覧 = 出庫情報シート.get_all_values()
    出庫先 = None
    for info_row in 出庫情報一覧[1:]:  # ヘッダーを除く
        if info_row[0] == 出庫ID:
            出庫先 = info_row[2]
            break

    if 出庫先 is None:
        print(f"出庫ID {出庫ID} に対応する出庫先が見つかりませんでした")
        return

    # 出庫先に応じて振り分けて記録
    振り分けデータ = [出庫ID, 商品名, 数量, 出庫先]
    if 出庫先 == "取引先への出荷":
        sheets["卸"].append_row(振り分けデータ)
    elif 出庫先 in ["通販", "店頭販売", "課税出荷（イベント等）", "持ち出し"]:
        sheets["店頭等"].append_row(振り分けデータ)


# --- メール送信用のルーティング ---
@app.route('/email', methods=['GET', 'POST'])
def send_email_page():
    items = get_available_items()
    if request.method == 'POST':
        raw_emails = request.form.get('emails', '')
        email_list = extract_valid_emails(raw_emails)
        body = create_email_body(items)
        if not email_list:
            flash('有効なメールアドレスが見つかりませんでした。', 'danger')
        else:
            for email in email_list:
                try:
                    send_email(body, email)
                    save_email_log_xlsx(email, '成功')
                except Exception as e:
                    save_email_log_xlsx(email, '失敗', str(e))
                    flash(f'{email} 宛ての送信に失敗しました: {str(e)}', 'danger')
            else:
                flash(f'{len(email_list)} 件のメールを送信しました。', 'success')
            return redirect(url_for('send_email_page'))
    return render_template('email_index.html', items=items)

@app.route('/logs')
def view_logs():
    logs = []
    header = []
    try:
        wb = load_workbook('email_log.xlsx')
        ws = wb.active
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                header = list(row)
            else:
                logs.append(list(row))
    except FileNotFoundError:
        header = ['送信先', '送信日時', '結果', 'エラーメッセージ']
        logs = []
    return render_template('logs.html', header=header, logs=logs)

# --- 以下、既存の出庫登録やCSVアップロード機能などを続けて追記 ---
# （この後ろに元の出庫管理ルーティングを貼り付けてください）


#プルダウン形式での出庫情報入力
def get_shukkosaki_options():
    sheet = gspread.authorize(ServiceAccountCredentials.from_json_keyfile_dict(
        json.load(open(os.environ['GOOGLE_CREDENTIALS_PATH'])),
        ['https://spreadsheets.google.com/feeds', 'https://www.googleapis.com/auth/drive']
    )).open('【開発用】シードル出庫台帳').worksheet('出庫先')
    return sheet.col_values(1)[1:]  # ヘッダーを除く

def get_product_options():
    sheet = gspread.authorize(ServiceAccountCredentials.from_json_keyfile_dict(
        json.load(open(os.environ['GOOGLE_CREDENTIALS_PATH'])),
        ['https://spreadsheets.google.com/feeds', 'https://www.googleapis.com/auth/drive']
    )).open('【開発用】シードル出庫台帳').worksheet('商品名')
    return sheet.col_values(1)[1:]

def get_staff_options():
    sheet = gspread.authorize(ServiceAccountCredentials.from_json_keyfile_dict(
        json.load(open(os.environ['GOOGLE_CREDENTIALS_PATH'])),
        ['https://spreadsheets.google.com/feeds', 'https://www.googleapis.com/auth/drive']
    )).open('【開発用】シードル出庫台帳').worksheet('スタッフ')
    return sheet.col_values(1)[1:]


# 新しい出庫IDを生成
def generate_unique_id(出庫情報シート):
    """日付ベースのユニークな出庫IDを生成する (例: 240521-001)"""
    today_str = datetime.now().strftime("%y%m%d")
    all_values = 出庫情報シート.get_all_values()
    # ヘッダー行を除き、今日の日付で始まるIDをフィルタリング
    all_ids = [row[0] for row in all_values[1:] if row and row[0].startswith(today_str)]
    counter = len(all_ids) + 1
    return f"{today_str}-{str(counter).zfill(3)}"

# SQLiteデータベースへの接続
def get_db():
    """SQLiteデータベースへの接続を取得する"""
    db = sqlite3.connect(app.config['DATABASE'])
    db.row_factory = sqlite3.Row
    return db

# SQLiteデータベースの初期化
def init_db():
    """データベーステーブルを初期化（存在しない場合のみ作成）"""
    with app.app_context():
        db = get_db()
        with db:
            db.execute('''
                CREATE TABLE IF NOT EXISTS alcohol_sales (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    date TEXT NOT NULL,
                    product_name TEXT NOT NULL,
                    sales_count INTEGER NOT NULL,
                    source_filename TEXT NOT NULL
                )
            ''')
            db.execute('''
                CREATE TABLE IF NOT EXISTS upload_log (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    filename TEXT NOT NULL,
                    file_hash TEXT UNIQUE NOT NULL,
                    uploaded_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            ''')
        db.close()

# 許可されたファイル拡張子かチェック
def allowed_file(filename):
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in app.config['ALLOWED_EXTENSIONS']

# ファイルのハッシュ値を計算（重複チェック用）
def calculate_file_hash(filepath):
    hasher = hashlib.md5()
    with open(filepath, 'rb') as f:
        hasher.update(f.read())
    return hasher.hexdigest()

# データベースを初期化
init_db()


# --- 3. ルーティング（画面表示と処理） ---

# === 既存の在庫管理機能 ===

@app.route('/register', methods=['GET', 'POST'])
def register():
    """手動での出庫情報登録ページ"""

    if request.method == 'POST':
        出庫日 = request.form['date']
        出庫先 = request.form['destination']
        担当者 = request.form['staff']
        取引先 = request.form.get('client', '')

        gc = gspread.authorize(ServiceAccountCredentials.from_json_keyfile_dict(
            json.load(open(os.environ['GOOGLE_CREDENTIALS_PATH'])),
            ['https://spreadsheets.google.com/feeds', 'https://www.googleapis.com/auth/drive']
        ))
        workbook = gc.open('【開発用】シードル出庫台帳')

        出庫情報シート = workbook.worksheet('出庫情報')
        出庫詳細シート = workbook.worksheet('出庫詳細')
        出庫詳細_卸シート = workbook.worksheet('出庫詳細（卸）')
        出庫詳細_直販シート = workbook.worksheet('出庫詳細（店頭、通販等）')

        出庫ID = generate_unique_id(出庫情報シート)

        # 出庫情報を書き込み
        出庫情報シート.append_row([出庫ID, 出庫日, 出庫先, 取引先, 担当者])

        # 商品データの収集
        details_to_append = []
        for i in range(1, 6):
            商品名 = request.form.get(f'item{i}')
            数量 = request.form.get(f'qty{i}')
            if 商品名 and 数量:
                details_to_append.append([出庫ID, 商品名, 数量])

        # 出庫詳細シート（共通）に書き込み
        if details_to_append:
            出庫詳細シート.append_rows(details_to_append, value_input_option='USER_ENTERED')

            # 出庫先に応じて詳細を別シートにも分岐保存
            if 出庫先 == '取引先への出荷':
                出庫詳細_卸シート.append_rows(details_to_append, value_input_option='USER_ENTERED')
            else:
                出庫詳細_直販シート.append_rows(details_to_append, value_input_option='USER_ENTERED')

        return render_template(
            'success.html',
            message="出庫情報を登録しました",
            redirect_url=url_for('register')
        )

    # GETリクエスト時：プルダウンの選択肢を取得してフォームに渡す
    def get_dropdown_values(sheet_name):
        sheet = gspread.authorize(ServiceAccountCredentials.from_json_keyfile_dict(
            json.load(open(os.environ['GOOGLE_CREDENTIALS_PATH'])),
            ['https://spreadsheets.google.com/feeds', 'https://www.googleapis.com/auth/drive']
        )).open('【開発用】シードル出庫台帳').worksheet(sheet_name)
        return sheet.col_values(1)[1:]  # ヘッダー行を除く

    shukkosaki_options = get_dropdown_values('出庫先')
    product_options = get_dropdown_values('商品名')
    staff_options = get_dropdown_values('スタッフ')

    return render_template(
        'register.html',
        shukkosaki_options=shukkosaki_options,
        product_options=product_options,
        staff_options=staff_options
    )


@app.route('/list')
def list_data():
    """出庫情報の一覧ページ"""
    _, 出庫情報シート, _ = connect_sheets() 
    出庫情報 = 出庫情報シート.get_all_values()
    return render_template('list.html', 出庫情報=出庫情報)


@app.route('/detail/<shukko_id>')
def detail(shukko_id):
    """出庫情報の詳細ページ"""
    _ , 出庫情報シート, 出庫詳細シート = connect_sheets()
    出庫情報リスト = 出庫情報シート.get_all_values()
    出庫情報 = next((row for row in 出庫情報リスト if row[0] == shukko_id), None)
    出庫詳細リスト = 出庫詳細シート.get_all_values()
    出庫詳細 = [row for row in 出庫詳細リスト if row[0] == shukko_id]
    return render_template('detail.html', 出庫情報=出庫情報, 出庫詳細=出庫詳細, 出庫ID=shukko_id)


@app.route('/edit/<shukko_id>', methods=['GET', 'POST'])
def edit(shukko_id):
    """出庫情報の編集ページ"""
    出庫情報シート, _, _ = connect_sheets()
    出庫情報リスト = 出庫情報シート.get_all_values()
    index = None
    出庫情報 = None
    for i, row in enumerate(出庫情報リスト):
        if row[0] == shukko_id:
            index = i + 1  # Sheetsは1始まり
            出庫情報 = row
            break
    
    if index is None:
        return "指定された出庫IDが見つかりません。", 404

    if request.method == 'POST':
        出庫日 = request.form['date']
        出庫先 = request.form['destination']
        取引先 = request.form['client']
        担当者 = request.form['staff']
        出庫情報シート.update(f'B{index}:E{index}', [[出庫日, 出庫先, 取引先, 担当者]])
        return render_template(
            'success.html',
            message="出庫情報を更新しました",
            redirect_url=url_for('detail', shukko_id=shukko_id)
        )
    return render_template('edit.html', 出庫ID=shukko_id, 出庫情報=出庫情報)


@app.route('/delete/<shukko_id>')
def delete_shukko(shukko_id):
    _ , 出庫情報シート, 出庫詳細シート = connect_sheets()

    # 出庫情報シートから対象行を削除
    cell = 出庫情報シート.find(shukko_id)
    if cell:
        出庫情報シート.delete_rows(cell.row)

    # 出庫詳細シートからも関連行を削除
    cells = 出庫詳細シート.findall(shukko_id)
    for c in reversed(cells):  # 後ろから削除しないとインデックスずれる
        出庫詳細シート.delete_rows(c.row)

    return redirect(url_for('list_data'))  # 一覧に戻る



@app.route('/edit-detail/<shukko_id>', methods=['GET', 'POST'])
def edit_detail(shukko_id):
    """出庫詳細の編集ページ"""
    _, _, 出庫詳細シート = connect_sheets()
    # 該当する出庫IDの行をすべて削除
    # gspreadには特定の条件で行を削除する簡単なAPIがないため、一度クリアして再追加するアプローチが一般的
    
    # ここは元のコードのロジックが複雑でgspreadのAPI制限に引っかかりやすいため、
    # より安全な「詳細ページに戻って、手動で修正してもらう」運用を推奨しますが、
    # 元のロジックをベースに簡易的な更新処理を実装します。

    if request.method == 'POST':
        # 既存の詳細を一旦クリア
        all_details = 出庫詳細シート.get_all_records() # 辞書形式で取得
        rows_to_delete = []
        for i, record in enumerate(all_details):
            if record.get('出庫ID') == shukko_id:
                rows_to_delete.append(i + 2) # ヘッダーを考慮した行番号

        # 後ろの行から削除しないとインデックスがずれる
        for row_num in sorted(rows_to_delete, reverse=True):
            出庫詳細シート.delete_rows(row_num)

        # フォームから新しい詳細を追加
        new_details = []
        for i in range(1, 11): # 登録できる商品数を増やすなど調整可能
             商品名 = request.form.get(f'item{i}')
             数量 = request.form.get(f'qty{i}')
             if 商品名 and 数量:
                 new_details.append([shukko_id, 商品名, 数量])
        
        if new_details:
            出庫詳細シート.append_rows(new_details, value_input_option='USER_ENTERED')

        return render_template(
            'success.html',
            message="出庫詳細を更新しました",
            redirect_url=url_for('detail', shukko_id=shukko_id)
        )
    
    # GETリクエストの場合、編集用のデータを渡す
    出庫詳細リスト = 出庫詳細シート.get_all_values()
    出庫詳細 = [row for row in 出庫詳細リスト if row[0] == shukko_id]
    return render_template('edit_detail.html', 出庫ID=shukko_id, 出庫詳細=出庫詳細)


@app.route('/edit-detail/<shukko_id>/<detail_id>', methods=['POST'])
def edit_shukko_detail(shukko_id, detail_id):
    new_name = request.form['product_name']
    new_qty = request.form['quantity']
    出庫詳細シート = connect_sheets()[1]

    # detail_idをキーにして該当行を探し更新
    cell = 出庫詳細シート.find(detail_id)
    if cell:
        row = cell.row
        出庫詳細シート.update_cell(row, 3, new_name)  # 商品名
        出庫詳細シート.update_cell(row, 4, new_qty)   # 数量

    return redirect(url_for('detail', 出庫ID=shukko_id))


@app.route('/delete-detail/<shukko_id>/<detail_id>')
def delete_detail(shukko_id, detail_id):
    出庫詳細シート = connect_sheets()[1]

    cell = 出庫詳細シート.find(detail_id)
    if cell:
        出庫詳細シート.delete_rows(cell.row)

    return redirect(url_for('detail', 出庫ID=shukko_id))



def process_and_store_csv(filepath, filename, file_hash):
    """CSVを解析し、「お酒類」のデータをDBとExcelに登録します。"""
    try:
        date_str = os.path.basename(filename).split('-')[0]
        sale_date = datetime.strptime(date_str, '%Y%m%d').strftime('%Y-%m-%d')
        print(f"\n--- Processing file: {filename} for date: {sale_date} ---")

        try:
            df = pd.read_csv(filepath, header=None, encoding='shift-jis')
        except Exception:
            df = pd.read_csv(filepath, header=None, encoding='utf-8')
        
        print(f"[DEBUG] CSV loaded. Shape: {df.shape}")
        if 1 in df.columns:
            unique_categories = df[1].astype(str).str.strip().unique()
            print(f"[DEBUG] Unique categories in column B: {unique_categories}")
        else:
            return jsonify({'error': 'CSVファイルにカテゴリ列（B列）が見つかりません。'}), 500

        df_filtered = df[df[1].astype(str).str.strip() == 'お酒類'].copy()
        if df_filtered.empty:
            return jsonify({'success': 'ファイルは処理されましたが、「お酒類」のデータは見つかりませんでした。'}), 200

        if 0 not in df_filtered.columns or 7 not in df_filtered.columns:
            return jsonify({'error': '必要な列（A列またはH列）が見つかりません。'}), 500

        result_df = df_filtered[[0, 7]].copy()
        result_df.columns = ['product_name', 'sales_count']
        result_df['sales_count'] = pd.to_numeric(result_df['sales_count'], errors='coerce').fillna(0).astype(int)
        result_df['date'] = sale_date
        result_df['source_filename'] = filename

        product_name_mapping = {
            "シードル辛口フル　2180円": "シードル辛口2025／フル",
            "シードル甘口ハーフ　1250円": "シードル甘口2025／ハーフ",
            "シードル辛口ハーフ　1250円": "シードル辛口2025／ハーフ",
            "シードル　低アルコール　2180円": "低アルコール2025／フル",
            "シードル甘口フル　2180円": "シードル甘口2025／フル",
            "洋梨スパークリング　フル　2600円": "洋梨／フル2025",
            "洋梨スパークリング　ハーフ　1500円": "洋梨／ハーフ2025",
            "ワインハーフボトル1500円": "ワイン／ハーフ2025",
            "ワインフルボトル2600円": "ワイン／フル2025",
            "シナノブレンド甘口　1250円": "シナノブレンド甘口2025",
            "シナノブレンド辛口　1250円": "シナノブレンド辛口2025",
            "シードル【フル】3本セット　6500円": "シードル3本セット2025／フル"
        }

        result_df['product_name'] = result_df['product_name'].astype(str).str.strip().map(product_name_mapping)

        # マッピングされなかった商品は除外
        result_df.dropna(subset=['product_name'], inplace=True)

        print("[DEBUG] Data to be inserted (first 5 rows):")
        print(result_df.head().to_string())

        conn = get_db()
        with conn:
            result_df.to_sql('alcohol_sales', conn, if_exists='append', index=False)
            conn.execute(
                'INSERT OR IGNORE INTO upload_log (filename, file_hash) VALUES (?, ?)',
                (filename, file_hash)
            )
        conn.close()

        # Google スプレッドシートへの追記
        try:
            sheet1, sheet2 = connect_sheets()
            existing_ids = sheet1.col_values(1)[1:]  # 出庫ID列（ヘッダー除く）
            today = datetime.strptime(sale_date, '%Y-%m-%d').strftime('%y%m%d')

            # 使用済み番号セット
            used_numbers = set()
            for id_ in existing_ids:
                match = re.match(rf'^{today}-(\d+)$', str(id_))
                if match:
                    used_numbers.add(int(match.group(1)))

            # 出庫IDジェネレーター定義
            def get_next_shukko_id(start=1):
                num = start
                while True:
                    if num not in used_numbers:
                        used_numbers.add(num)
                        yield f'{today}-{num:03d}'
                    num += 1

            id_generator = get_next_shukko_id()
            shukko_ids = []

            for _, row in result_df.iterrows():
                shukko_id = next(id_generator)
                shukko_ids.append(shukko_id)

            # Google Sheets に書き込む
            write_to_google_sheets(result_df, shukko_ids)
            print("[INFO] Googleスプレッドシートに出庫データを追加しました。")

        except Exception as e:
            print(f"[ERROR] Google Sheetsへの書き込みに失敗しました: {e}")

        print("--- Processing finished successfully ---")
        return jsonify({'success': 'ファイルが正常に処理され、データベースおよびスプレッドシートに登録されました。'}), 200

    except Exception as e:
        print(f"!!! An error occurred: {e} !!!")
        traceback.print_exc()
        return jsonify({'error': f'CSV処理中に予期せぬエラーが発生しました: {e}'}), 500


def write_to_google_sheets(result_df, shukko_ids):
    try:
        # シートに接続
        sheet1, sheet2 = connect_sheets()

        for i, (_, row) in enumerate(result_df.iterrows()):
            shukko_id = shukko_ids[i]

            # 出庫情報シート：出庫ID・日付・出庫先・取引先・担当者
            sheet1.append_row([
                shukko_id,
                row['date'],
                '店頭販売',
                '',
                '北沢'
            ])

            # 出庫詳細シート：出庫ID・商品名・数量
            sheet2.append_row([
                shukko_id,
                row['product_name'],
                row['sales_count']
            ])

        print("[INFO] Googleスプレッドシートにデータを追加しました。")
    except Exception as e:
        print(f"[ERROR] Googleスプレッドシートへの書き込みに失敗しました: {e}")




# === 新しいCSVアップロード機能 ===

@app.route('/')
def index():
    """CSVアップロードページ（トップページ）"""
    return render_template('index.html')

@app.route('/upload', methods=['POST'])
def upload_file():
    """CSVファイルのアップロード処理"""
    if 'file' not in request.files:
        return jsonify({'error': 'ファイルがありません'}), 400
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'ファイルが選択されていません'}), 400

    if file and allowed_file(file.filename):
        filename = secure_filename(file.filename)
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(filepath)

        try:
            file_hash = calculate_file_hash(filepath)
            db = get_db()
            cur = db.execute('SELECT * FROM upload_log WHERE file_hash = ?', (file_hash,))
            existing = cur.fetchone()
            db.close()

            if existing:
                return jsonify({
                    'status': 'confirm',
                    'message': '同じ内容のファイルが既にアップロードされています。上書きして処理を続行しますか？ (既存のデータは削除されます)',
                    'filename': filename,
                    'file_hash': file_hash
                })

            return process_and_store_csv(filepath, filename, file_hash)

        except Exception as e:
            print(f"!!! An error occurred during upload: {e} !!!")
            traceback.print_exc()
            if os.path.exists(filepath):
                os.remove(filepath)
            return jsonify({'error': f'ファイル処理中にエラーが発生しました: {e}'}), 500
    else:
        return jsonify({'error': '許可されていないファイル形式です'}), 400

@app.route('/confirm_upload', methods=['POST'])
def confirm_upload():
    """重複確認後の再アップロード処理"""
    filename = request.form['filename']
    file_hash = request.form['file_hash']
    filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    
    try:
        db = get_db()
        with db:
            db.execute('DELETE FROM alcohol_sales WHERE source_filename = ?', (filename,))
        db.close()
        return process_and_store_csv(filepath, filename, file_hash)
    except Exception as e:
        print(f"!!! An error occurred during re-processing: {e} !!!")
        traceback.print_exc()
        return jsonify({'error': f'再処理中にエラーが発生しました: {e}'}), 500

        return jsonify({'error': f'CSV処理中に予期せぬエラーが発生しました: {e}'}), 500


# SQLiteのデータを表示・管理するページ
@app.route('/data')
def show_data():
    db = get_db()
    entries = db.execute('SELECT id, date, product_name, sales_count, source_filename FROM alcohol_sales ORDER BY date DESC, id DESC').fetchall()
    db.close()
    return render_template('data.html', entries=entries)

@app.route('/delete/<int:id>', methods=['POST'])
def delete_entry(id):
    db = get_db()
    with db:
        db.execute('DELETE FROM alcohol_sales WHERE id = ?', (id,))
    db.close()
    return redirect(url_for('show_data'))

@app.route('/delete_all', methods=['POST'])
def delete_all_entries():
    db = get_db()
    with db:
        db.execute('DELETE FROM alcohol_sales')
        db.execute('DELETE FROM upload_log')
    db.close()
    return redirect(url_for('show_data'))


# --- 4. アプリケーションの実行 ---
if __name__ == '__main__':
    app.run(debug=True)
